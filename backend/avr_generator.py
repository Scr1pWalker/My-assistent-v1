"""
АВР генератор — форма Р-1
Приложение 50 к приказу МФ РК от 20.12.2012 № 562

Динамические строки: сколько услуг — столько строк (без ограничений).
Пустые строки не создаются.
"""

import os, copy, subprocess, tempfile, re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "avr_template.xlsx")

# ── Тонкая граница (как в оригинале) ─────────
THIN = Side(style='thin')

def border(**sides):
    return Border(
        top=    sides.get('top'),
        bottom= sides.get('bottom'),
        left=   sides.get('left'),
        right=  sides.get('right'),
    )

# Точные границы для каждой группы колонок строки услуги
# (скопировано из оригинала)
COL_BORDERS = {
    'A':  border(top=THIN, bottom=THIN, left=THIN, right=THIN),   # №
    'B':  border(top=THIN, bottom=THIN,             right=THIN),
    # C..O — Наименование
    'C':  border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    **{c: border(top=THIN, bottom=THIN) for c in ['D','E','F','G','H','I','J','K','L','M','N','O']},
    'O':  border(top=THIN, bottom=THIN,             right=THIN),
    # P..T — Дата выполнения
    'P':  border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    **{c: border(top=THIN, bottom=THIN) for c in ['Q','R','S','T']},
    'T':  border(top=THIN, bottom=THIN,             right=THIN),
    # U..AC — Сведения об отчёте
    'U':  border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    **{c: border(top=THIN, bottom=THIN) for c in ['V','W','X','Y','Z','AA','AB','AC']},
    'AC': border(top=THIN, bottom=THIN,             right=THIN),
    # AD..AF — Единица измерения
    'AD': border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    'AE': border(top=THIN, bottom=THIN),
    'AF': border(top=THIN, bottom=THIN,             right=THIN),
    # AG..AK — Количество
    'AG': border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    **{c: border(top=THIN, bottom=THIN) for c in ['AH','AI','AJ','AK']},
    'AK': border(top=THIN, bottom=THIN,             right=THIN),
    # AL..AQ — Цена за единицу
    'AL': border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    **{c: border(top=THIN, bottom=THIN) for c in ['AM','AN','AO','AP','AQ']},
    'AQ': border(top=THIN, bottom=THIN,             right=THIN),
    # AR..AW — Стоимость
    'AR': border(top=THIN, bottom=THIN, left=THIN,  right=THIN),
    **{c: border(top=THIN, bottom=THIN) for c in ['AS','AT','AU','AV','AW']},
    'AW': border(top=THIN, bottom=THIN,             right=THIN),
}

# Merge ranges для одной строки услуги (относительно строки)
ROW_MERGES = [
    ('A','B'),    # №
    ('C','O'),    # Наименование
    ('P','T'),    # Дата выполнения
    ('U','AC'),   # Сведения об отчёте
    ('AD','AF'),  # Единица измерения
    ('AG','AK'),  # Количество
    ('AL','AQ'),  # Цена за единицу
    ('AR','AW'),  # Стоимость
]

def col_num(col_letter):
    """A→1, B→2, AW→49"""
    n = 0
    for c in col_letter:
        n = n * 26 + (ord(c) - ord('A') + 1)
    return n


def set_cell(ws, addr, value, bold=False, size=8, halign='center', valign='center', wrap=True, fmt=None):
    cell = ws[addr]
    cell.value = value
    cell.font = Font(name='Arial', size=size, bold=bold)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if fmt:
        cell.number_format = fmt


def write_service_row(ws, row_num, idx, svc):
    """Записывает одну строку услуги с правильными границами и merge"""

    # 1. Снимаем все старые merge в этой строке (если есть)
    to_remove = [m for m in ws.merged_cells.ranges
                 if m.min_row <= row_num <= m.max_row
                 and m.min_row == row_num]
    for m in to_remove:
        ws.merged_cells.remove(m)

    # 2. Устанавливаем границы всем ячейкам строки
    for col_letter, brd in COL_BORDERS.items():
        col_idx = col_num(col_letter)
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = brd

    # 3. Merge ячейки
    for start_col, end_col in ROW_MERGES:
        ws.merge_cells(f'{start_col}{row_num}:{end_col}{row_num}')

    # 4. Высота строки
    ws.row_dimensions[row_num].height = 21.75

    # 5. Данные
    qty   = float(svc.get('qty', 1) or 1)
    price = float(svc.get('price', 0) or 0)

    set_cell(ws, f'A{row_num}',  str(idx + 1),              halign='center')
    set_cell(ws, f'C{row_num}',  svc.get('name', ''),       halign='left')
    set_cell(ws, f'AD{row_num}', svc.get('unit', 'Услуга'), halign='center')
    set_cell(ws, f'AG{row_num}', qty,                        halign='right', fmt='#,##0.##')
    set_cell(ws, f'AL{row_num}', price,                      halign='right', fmt='#,##0.00')

    # Сумма — формула Excel
    ws[f'AR{row_num}'] = f'=AL{row_num}*AG{row_num}'
    ws[f'AR{row_num}'].font      = Font(name='Arial', size=8)
    ws[f'AR{row_num}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'AR{row_num}'].number_format = '#,##0.00'

    # Дата выполнения
    wd = svc.get('workDate', '')
    if wd:
        try:
            wdt = datetime.strptime(wd, '%d.%m.%Y')
            ws[f'P{row_num}'] = wdt
            ws[f'P{row_num}'].number_format = 'DD.MM.YYYY'
            ws[f'P{row_num}'].font      = Font(name='Arial', size=8)
            ws[f'P{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        except:
            set_cell(ws, f'P{row_num}', wd, halign='center')

    # Сведения об отчёте
    ri = svc.get('reportInfo', '')
    if ri:
        set_cell(ws, f'U{row_num}', ri, halign='center')


def shift_rows_down(ws, from_row, count):
    """
    Сдвигает строки вниз начиная с from_row на count позиций.
    Используется когда услуг больше 4 (размер шаблона).
    """
    ws.insert_rows(from_row, amount=count)


def generate_avr_xlsx(data: dict, output_path: str):
    services = data.get('services', [])
    n = len(services)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    buyer  = data.get('buyer', {})
    seller = data.get('seller', {})

    # ── Стороны ───────────────────────────────
    buyer_str  = ', '.join(filter(None, [buyer.get('companyName',''),  buyer.get('address','')]))
    seller_str = ', '.join(filter(None, [seller.get('companyName',''), seller.get('address','')]))

    set_cell(ws, 'E9',   buyer_str,              bold=True, size=9, halign='center')
    set_cell(ws, 'AQ9',  buyer.get('bin',''),    bold=True, size=9, halign='center')
    set_cell(ws, 'E11',  seller_str,             bold=True, size=9, halign='center')
    set_cell(ws, 'AQ11', seller.get('bin',''),   bold=True, size=9, halign='center')

    # ── Договор ───────────────────────────────
    set_cell(ws, 'F13', data.get('contract',''), halign='left')

    # ── Номер и дата ──────────────────────────
    set_cell(ws, 'AH15', str(data.get('number','1')), bold=True, halign='center')
    date_str = data.get('date', '')
    try:
        dt = datetime.strptime(date_str, '%d.%m.%Y')
        ws['AM15'] = dt
        ws['AM15'].number_format = 'DD.MM.YYYY'
        ws['AM15'].font = Font(name='Arial', size=8, bold=True)
        ws['AM15'].alignment = Alignment(horizontal='center', vertical='center')
    except:
        set_cell(ws, 'AM15', date_str, bold=True, halign='center')

    # ── Услуги ────────────────────────────────
    # Шаблон имеет 4 строки (20-23) и строки после (24+).
    # Нам нужно n строк. Если n < 4 — удаляем лишние.
    # Если n > 4 — вставляем новые.

    FIRST_SVC_ROW = 20
    TEMPLATE_SVC_ROWS = 4  # в шаблоне строк 20,21,22,23
    ITOGO_ROW = 24         # строка "Итого" в шаблоне

    if n < TEMPLATE_SVC_ROWS:
        # Удаляем лишние строки снизу (строки n+20 .. 23)
        rows_to_delete = TEMPLATE_SVC_ROWS - n
        delete_from = FIRST_SVC_ROW + n
        ws.delete_rows(delete_from, rows_to_delete)
        # Итого теперь на строке FIRST_SVC_ROW + n
        itogo_row = FIRST_SVC_ROW + n

    elif n > TEMPLATE_SVC_ROWS:
        # Вставляем дополнительные строки перед "Итого"
        extra = n - TEMPLATE_SVC_ROWS
        ws.insert_rows(ITOGO_ROW, extra)
        itogo_row = ITOGO_ROW + extra
    else:
        itogo_row = ITOGO_ROW

    # Записываем строки услуг
    for i, svc in enumerate(services):
        row_num = FIRST_SVC_ROW + i
        write_service_row(ws, row_num, i, svc)


    # ── Строка Итого ──────────────────────────
    # Снимаем ВСЕ merge в строке итого
    to_remove = [m for m in list(ws.merged_cells.ranges)
                 if m.min_row == itogo_row]
    for m in to_remove:
        ws.merged_cells.remove(m)

    # Устанавливаем границы
    for col_idx in range(1, 50):
        col_l = get_column_letter(col_idx)
        ws.cell(row=itogo_row, column=col_idx).border = COL_BORDERS.get(
            col_l, border(top=THIN, bottom=THIN))

    # Записываем значения ДО merge (иначе MergedCell — read-only)
    first_r = FIRST_SVC_ROW
    last_r  = FIRST_SVC_ROW + n - 1

    ws.cell(row=itogo_row, column=32).value = 'Итого'   # AF
    ws.cell(row=itogo_row, column=32).font = Font(name='Arial', size=8)
    ws.cell(row=itogo_row, column=32).alignment = Alignment(horizontal='right', vertical='center')

    ws.cell(row=itogo_row, column=38).value = 'х'       # AL
    ws.cell(row=itogo_row, column=38).font = Font(name='Arial', size=8)
    ws.cell(row=itogo_row, column=38).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=itogo_row, column=33).value = f'=SUM(AG{first_r}:AG{last_r})'  # AG
    ws.cell(row=itogo_row, column=33).font = Font(name='Arial', size=8)
    ws.cell(row=itogo_row, column=33).alignment = Alignment(horizontal='right', vertical='center')

    ws.cell(row=itogo_row, column=44).value = f'=SUM(AR{first_r}:AR{last_r})'  # AR
    ws.cell(row=itogo_row, column=44).font = Font(name='Arial', size=8)
    ws.cell(row=itogo_row, column=44).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=itogo_row, column=44).number_format = '#,##0.00'

    # Merge ПОСЛЕ записи значений
    ws.merge_cells(f'A{itogo_row}:AE{itogo_row}')
    ws.merge_cells(f'AF{itogo_row}:AK{itogo_row}')
    ws.merge_cells(f'AL{itogo_row}:AQ{itogo_row}')
    ws.merge_cells(f'AR{itogo_row}:AW{itogo_row}')


    # ── Подписи ───────────────────────────────
    # Строки подписей смещаются вместе с вставкой/удалением строк услуг
    shift = n - TEMPLATE_SVC_ROWS  # может быть отрицательным

    signer_row = 34 + shift
    label_row  = 35 + shift
    mp_row     = 37 + shift
    mp2_row    = 39 + shift

    signer = seller.get('signerName', '')
    if signer:
        set_cell(ws, f'S{signer_row}', signer, halign='left')

    # Дата подписания
    try:
        dt = datetime.strptime(date_str, '%d.%m.%Y')
        ws[f'AL{mp_row}'] = dt
        ws[f'AL{mp_row}'].number_format = 'DD.MM.YYYY'
        ws[f'AL{mp_row}'].font = Font(name='Arial', size=8)
        ws[f'AL{mp_row}'].alignment = Alignment(horizontal='center', vertical='center')
    except:
        set_cell(ws, f'AL{mp_row}', date_str, halign='center')

    wb.save(output_path)


def xlsx_to_pdf(xlsx_path: str, pdf_path: str):
    out_dir = os.path.dirname(pdf_path)
    subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'pdf',
         '--outdir', out_dir, xlsx_path],
        capture_output=True, text=True, timeout=30
    )
    generated = xlsx_path.replace('.xlsx', '.pdf')
    if os.path.exists(generated) and generated != pdf_path:
        os.rename(generated, pdf_path)


def generate_avr_pdf(data: dict) -> bytes:
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, f"avr_{data.get('number','1')}.xlsx")
        pdf_path  = os.path.join(tmpdir, f"avr_{data.get('number','1')}.pdf")
        generate_avr_xlsx(data, xlsx_path)
        xlsx_to_pdf(xlsx_path, pdf_path)
        with open(pdf_path, 'rb') as f:
            return f.read()


if __name__ == '__main__':
    # Тест с 2 услугами
    test2 = {
        'number': '5', 'date': '19.06.2026',
        'contract': '№ 1 от 01.01.2026',
        'buyer':  {'companyName': 'ТОО "Тест"', 'bin': '123456789012', 'address': 'г. Алматы'},
        'seller': {'companyName': 'ИП Иванов', 'bin': '950121350285', 'address': 'г. Астана', 'signerName': 'Иванов И.И.'},
        'services': [
            {'name': 'Разработка приложения', 'unit': 'Услуга', 'qty': 1, 'price': 350000, 'workDate': '19.06.2026'},
            {'name': 'Техподдержка', 'unit': 'Месяц', 'qty': 3, 'price': 50000, 'workDate': ''},
        ]
    }
    pdf = generate_avr_pdf(test2)
    with open('/home/claude/test_avr_2rows.pdf', 'wb') as f: f.write(pdf)
    print(f"2 services: {len(pdf)} bytes ✓")

    # Тест с 6 услугами
    test6 = {**test2, 'number': '6'}
    test6['services'] = [
        {'name': f'Услуга {i+1}', 'unit': 'Услуга', 'qty': i+1, 'price': 50000*(i+1), 'workDate': ''}
        for i in range(6)
    ]
    pdf = generate_avr_pdf(test6)
    with open('/home/claude/test_avr_6rows.pdf', 'wb') as f: f.write(pdf)
    print(f"6 services: {len(pdf)} bytes ✓")
